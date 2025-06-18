import os

def count_pages(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        return count_pdf_pages(file_path)
    elif ext == ".docx":
        return count_docx_pages(file_path)
    elif ext == ".txt":
        return count_txt_pages(file_path)
    elif ext in [".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tiff"]:
        return 1  # каждое изображение — одна страница
    elif ext in [".xlsx", ".xls"]:
        return count_excel_pages(file_path)
    elif ext in [".pptx", ".ppt"]:
        return count_ppt_pages(file_path)
    elif ext == ".odt":
        return count_odt_pages(file_path)
    elif ext == ".ods":
        return count_ods_pages(file_path)
    elif ext == ".odp":
        return count_odp_pages(file_path)
    else:
        raise ValueError(f"Неизвестный формат файла: {ext}")


def count_pdf_pages(path):
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(path)
        return len(reader.pages)
    except Exception as e:
        print(f"Ошибка при чтении PDF: {e}")
        return 1

def count_docx_pages(path):
    try:
        import docx
        doc = docx.Document(path)
        word_count = sum(len(p.text.split()) for p in doc.paragraphs)
        return max(1, word_count // 500 + 1)
    except Exception as e:
        print(f"Ошибка при чтении DOCX: {e}")
        return 1

def count_txt_pages(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
        return max(1, len(lines) // 50 + 1)
    except Exception as e:
        print(f"Ошибка при чтении TXT: {e}")
        return 1

def count_excel_pages(path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        sheet_count = len(wb.sheetnames)
        row_count = sum(sheet.max_row for sheet in wb.worksheets)
        wb.close()
        return max(1, row_count // 50 + sheet_count)
    except Exception as e:
        print(f"Ошибка при чтении Excel: {e}")
        return 1

def count_ppt_pages(path):
    try:
        from pptx import Presentation
        prs = Presentation(path)
        return len(prs.slides)
    except Exception as e:
        print(f"Ошибка при чтении PowerPoint: {e}")
        return 1

def count_odt_pages(path):
    try:
        import ezodf
        doc = ezodf.opendoc(path)
        text = ""
        for para in doc.body:
            text += para.text
        word_count = len(text.split())
        return max(1, word_count // 500 + 1)
    except Exception as e:
        print(f"Ошибка при чтении ODT: {e}")
        return 1

def count_ods_pages(path):
    try:
        import ezodf
        doc = ezodf.opendoc(path)
        sheet_count = len(doc.sheets)
        row_count = sum(len(sheet.rows()) for sheet in doc.sheets)
        return max(1, row_count // 50 + sheet_count)
    except Exception as e:
        print(f"Ошибка при чтении ODS: {e}")
        return 1

def count_odp_pages(path):
    try:
        # ODP часто сложен для чтения, считаем как PPT
        return count_ppt_pages(path)
    except Exception as e:
        print(f"Ошибка при чтении ODP: {e}")
        return 1
